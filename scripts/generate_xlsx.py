#!/usr/bin/env python3
"""
测试用例 Excel 生成脚本
从 JSON 数据生成格式化的 .xlsx 测试用例文件

用法:
  # 方式1: 文件路径（本地开发）
  python generate_xlsx.py config.json output.xlsx

  # 方式2: JSON内容（企业系统集成）
  python generate_xlsx.py '{"cases": [...], "title": "..."}' output.xlsx

JSON 结构:
{
  "cases": [
    {
      "用例编号": "MOD001-FUN001-CASE001",
      "模块名称": "用户中心",
      "功能点": "手机号+验证码登录",
      "用例标题": "正确手机号+正确验证码登录成功",
      "前置条件": "1. 用户已注册\\n2. 系统网络正常",
      "测试步骤": "1. 打开登录页面\\n2. 输入手机号...",
      "预期结果": "登录成功，跳转到首页",
      "优先级": "高",
      "用例类型": "功能"
    }
  ],
  "extra_columns": ["测试环境", "执行人"],  // 可选
  "title": "XXX功能测试用例"               // 可选
}

企业系统集成示例:
  // Node.js
  const { spawn } = require('child_process');
  const config = { cases: [...], title: '测试用例' };

  spawn('python', ['generate_xlsx.py', JSON.stringify(config), 'output.xlsx']);

  // Python
  import subprocess
  config = {"cases": [...], "title": "测试用例"}
  subprocess.run(['python', 'generate_xlsx.py', json.dumps(config), 'output.xlsx'])
"""

import json
import sys
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is not installed. Install it with: pip install openpyxl")
    sys.exit(1)


# ── 样式常量 ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10)
EVEN_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ODD_ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="163B5E"),
    right=Side(style="thin", color="163B5E"),
    top=Side(style="thin", color="163B5E"),
    bottom=Side(style="thin", color="163B5E"),
)

# 优先级颜色
PRIORITY_COLORS = {
    "高": Font(name="微软雅黑", size=10, bold=True, color="C0392B"),
    "中": Font(name="微软雅黑", size=10, color="E67E22"),
    "低": Font(name="微软雅黑", size=10, color="7F8C8D"),
}

# 用例类型颜色
TYPE_COLORS = {
    "功能": Font(name="微软雅黑", size=10, color="2E86C1"),
    "边界": Font(name="微软雅黑", size=10, color="D4AC0D"),
    "异常": Font(name="微软雅黑", size=10, color="C0392B"),
}

# 核心字段顺序
CORE_COLUMNS = [
    "用例编号", "模块名称", "功能点", "用例标题",
    "前置条件", "测试步骤", "预期结果", "优先级", "用例类型"
]

# 列宽配置 (字符数)
COLUMN_WIDTHS = {
    "用例编号": 26,
    "模块名称": 14,
    "功能点": 22,
    "用例标题": 36,
    "前置条件": 32,
    "测试步骤": 42,
    "预期结果": 36,
    "优先级": 10,
    "用例类型": 10,
}
DEFAULT_EXTRA_WIDTH = 18


def create_workbook(data: dict, output_path: str):
    """生成格式化的测试用例 Excel 文件"""
    cases = data.get("cases", [])
    extra_columns = data.get("extra_columns", [])
    title = data.get("title", "测试用例")

    if not cases:
        print("WARNING: No test cases found in input data.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    # ── 构建列头 ──
    all_columns = CORE_COLUMNS + [c for c in extra_columns if c not in CORE_COLUMNS]

    # ── 写入表头 ──
    for col_idx, col_name in enumerate(all_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER

    # ── 写入数据 ──
    for row_idx, case in enumerate(cases, 2):
        row_fill = EVEN_ROW_FILL if row_idx % 2 == 0 else ODD_ROW_FILL

        for col_idx, col_name in enumerate(all_columns, 1):
            value = case.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
                horizontal="center" if col_name in ("用例编号", "优先级", "用例类型") else "left"
            )

            # 特殊字体
            if col_name == "优先级" and value in PRIORITY_COLORS:
                cell.font = PRIORITY_COLORS[value]
            elif col_name == "用例类型" and value in TYPE_COLORS:
                cell.font = TYPE_COLORS[value]
            else:
                cell.font = DATA_FONT

    # ── 设置列宽 ──
    for col_idx, col_name in enumerate(all_columns, 1):
        width = COLUMN_WIDTHS.get(col_name, DEFAULT_EXTRA_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 冻结首行 ──
    ws.freeze_panes = "A2"

    # ── 设置行高 ──
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(cases) + 2):
        ws.row_dimensions[row_idx].height = 45

    # ── 自动筛选 ──
    last_col = get_column_letter(len(all_columns))
    ws.auto_filter.ref = f"A1:{last_col}{len(cases) + 1}"

    # ── 保存 ──
    wb.save(output_path)
    print(f"OK: Generated {len(cases)} test cases -> {output_path}")


def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_xlsx.py <input.json> <output.xlsx>")
        print("       python generate_xlsx.py <json_content> <output.xlsx>")
        print()
        print("支持两种调用方式:")
        print("  1. 文件路径: python generate_xlsx.py config.json output.xlsx")
        print("  2. JSON内容: python generate_xlsx.py '{\"cases\": [...]}' output.xlsx")
        sys.exit(1)

    param1 = sys.argv[1]
    output_path = sys.argv[2]

    # 判断第一个参数是文件路径还是JSON内容
    if param1.startswith('{') or param1.startswith('['):
        # 企业系统直接传递JSON内容
        try:
            data = json.loads(param1)
            print(f"[INFO] Loaded {len(data.get('cases', []))} test cases from JSON content", file=sys.stderr)
        except json.JSONDecodeError as e:
            print(f"ERROR: Invalid JSON content: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        # 本地文件路径
        input_path = param1
        if not os.path.exists(input_path):
            print(f"ERROR: Input file not found: {input_path}", file=sys.stderr)
            sys.exit(1)

        try:
            with open(input_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            print(f"[INFO] Loaded {len(data.get('cases', []))} test cases from file: {input_path}", file=sys.stderr)
        except json.JSONDecodeError as e:
            print(f"ERROR: Invalid JSON file: {e}", file=sys.stderr)
            sys.exit(1)
        except Exception as e:
            print(f"ERROR: Failed to read file: {e}", file=sys.stderr)
            sys.exit(1)

    create_workbook(data, output_path)


if __name__ == "__main__":
    main()
